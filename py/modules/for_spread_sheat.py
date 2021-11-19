import openpyxl

def process_workbook(filename):
    wb=openpyxl.load_workbook(filename)
    sheet=wb['Sheet1']
    cell=sheet.cell(1,1)
    cell=sheet.cell(1,1)

    for row in range(2,sheet.max_row + 1):
        cell=sheet.cell(row,3)
        correct_price=cell.value*0.9
        correct_price_cell=sheet.cell(row,4)
        correct_price_cell.value=correct_price

    wb.save(filename)